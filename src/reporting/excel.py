# Copyright 2025 Mission Critical Email LLC. All rights reserved.
# Licensed under the MIT License. See LICENSE file in the project root.
#
# DISCLAIMER:
# This software is provided "AS IS" without warranty of any kind, either express
# or implied, including but not limited to the implied warranties of
# merchantability and fitness for a particular purpose. Use at your own risk.
# In no event shall Mission Critical Email LLC be liable for any damages
# whatsoever arising out of the use of or inability to use this software.

"""Excel report generation for monthly billing summaries.

Generates comprehensive Excel reports with:
- Summary sheet with totals by customer
- Detailed usage sheet
- Non-billable usage tracking
- Month-over-month comparisons
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..database.queries import QueryHelper
from ..database.models import Customer, Domain, MonthlyHighwater

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Generates Excel reports for billing data."""

    def __init__(self, query_helper: QueryHelper, qbo_client=None):
        """Initialize report generator.

        Args:
            query_helper: Database query helper
            qbo_client: Optional QuickBooks client for fetching current prices
        """
        self.query_helper = query_helper
        self.qbo_client = qbo_client
        self._price_cache = {}  # Cache prices to avoid repeated API calls

        # Styling
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.total_font = Font(bold=True, size=11)
        self.total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _get_item_price(self, qbo_item_id: str) -> float:
        """Get current price for a QuickBooks item.

        Args:
            qbo_item_id: QuickBooks item ID

        Returns:
            Current unit price from QuickBooks, or 0.0 if unavailable
        """
        # Check cache first
        if qbo_item_id in self._price_cache:
            return self._price_cache[qbo_item_id]

        # If no QBO client, return 0.0
        if not self.qbo_client:
            return 0.0

        try:
            # Fetch item from QuickBooks
            item = self.qbo_client.get_item_by_id(qbo_item_id)
            if item and hasattr(item, 'UnitPrice'):
                price = float(item.UnitPrice or 0)
                self._price_cache[qbo_item_id] = price
                return price
        except Exception as e:
            logger.warning(f"Could not fetch price for item {qbo_item_id}: {e}")

        return 0.0

    def generate_monthly_report(self, year: int, month: int,
                               output_path: Optional[str] = None) -> str:
        """Generate comprehensive monthly billing report.

        Args:
            year: Year
            month: Month
            output_path: Optional output file path

        Returns:
            Path to generated Excel file
        """
        logger.info(f"Generating monthly report for {year}-{month:02d}")

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Generate sheets
        self._create_summary_sheet(wb, year, month)
        self._create_detailed_usage_sheet(wb, year, month)
        self._create_customer_breakdown_sheet(wb, year, month)
        # Non-billable usage tab removed - reconciliation process eliminates unmapped items

        # Determine output path
        if output_path is None:
            from ..config import get_config
            config = get_config()
            data_dir = config.data_dir
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = str(data_dir / f"billing_report_{year}_{month:02d}_{timestamp}.xlsx")

        # Save workbook
        wb.save(output_path)
        logger.info(f"Report saved to {output_path}")

        return output_path

    def _create_summary_sheet(self, wb: Workbook, year: int, month: int) -> None:
        """Create summary sheet with high-level statistics.

        Args:
            wb: Workbook
            year: Year
            month: Month
        """
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws['A1'] = f"Zimbra Billing Report - {self._get_month_name(month)} {year}"
        ws['A1'].font = Font(bold=True, size=14)

        # Report metadata
        row = 3
        ws[f'A{row}'] = "Report Generated:"
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row += 1
        ws[f'A{row}'] = "Billing Period:"
        ws[f'B{row}'] = f"{year}-{month:02d}"

        # Get summary data
        highwater_data = self.query_helper.get_highwater_for_month(year, month, billable_only=True)

        # Calculate totals
        from collections import defaultdict
        customer_totals = defaultdict(lambda: {'accounts': 0, 'amount': 0.0})

        for hw in highwater_data:
            domain = self.query_helper.session.query(Domain).get(hw.domain_id)
            if not domain:
                continue

            customer = self.query_helper.session.query(Customer).get(domain.customer_id)
            if not customer:
                continue

            cos_mapping = self.query_helper.get_cos_mapping_by_id(hw.cos_id)
            if not cos_mapping:
                continue

            # Fetch current price from QuickBooks
            unit_price = self._get_item_price(cos_mapping.qbo_item_id)
            amount = hw.highwater_count * unit_price

            customer_totals[customer.customer_name]['accounts'] += hw.highwater_count
            customer_totals[customer.customer_name]['amount'] += amount

        # Summary statistics
        row = 6
        ws[f'A{row}'] = "Summary Statistics"
        ws[f'A{row}'].font = self.total_font
        row += 1

        total_customers = len(customer_totals)
        total_accounts = sum(c['accounts'] for c in customer_totals.values())
        total_amount = sum(c['amount'] for c in customer_totals.values())

        ws[f'A{row}'] = "Total Customers:"
        customers_cell = ws[f'B{row}']
        customers_cell.value = total_customers
        customers_cell.alignment = Alignment(horizontal='right')
        row += 1
        ws[f'A{row}'] = "Total Accounts:"
        accounts_cell = ws[f'B{row}']
        accounts_cell.value = total_accounts
        accounts_cell.alignment = Alignment(horizontal='right')
        row += 1
        ws[f'A{row}'] = "Total Revenue:"
        revenue_cell = ws[f'B{row}']
        revenue_cell.value = f"${total_amount:,.2f}"
        revenue_cell.font = self.total_font
        revenue_cell.alignment = Alignment(horizontal='right')

        # Customer summary table
        row += 2
        ws[f'A{row}'] = "Customer"
        ws[f'B{row}'] = "Total Accounts"
        ws[f'C{row}'] = "Total Amount"

        # Style header row
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{row}']
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')

        row += 1

        # Add customer rows
        for customer_name in sorted(customer_totals.keys()):
            data = customer_totals[customer_name]
            ws[f'A{row}'] = customer_name
            # Center the account count
            accounts_cell = ws[f'B{row}']
            accounts_cell.value = data['accounts']
            accounts_cell.alignment = Alignment(horizontal='center')
            # Right-justify dollar amount
            amount_cell = ws[f'C{row}']
            amount_cell.value = f"${data['amount']:,.2f}"
            amount_cell.alignment = Alignment(horizontal='right')
            row += 1

        # Totals row
        ws[f'A{row}'] = "TOTAL"
        # Center total accounts
        total_accounts_cell = ws[f'B{row}']
        total_accounts_cell.value = total_accounts
        total_accounts_cell.alignment = Alignment(horizontal='center')
        # Right-justify total amount
        total_amount_cell = ws[f'C{row}']
        total_amount_cell.value = f"${total_amount:,.2f}"
        total_amount_cell.alignment = Alignment(horizontal='right')

        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = self.total_font
            ws[f'{col}{row}'].fill = self.total_fill

        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    def _create_detailed_usage_sheet(self, wb: Workbook, year: int, month: int) -> None:
        """Create detailed usage sheet with all billable line items.

        Args:
            wb: Workbook
            year: Year
            month: Month
        """
        ws = wb.create_sheet("Detailed Usage")

        # Headers
        headers = ['Customer', 'Domain', 'CoS', 'Quota (GB)', 'Accounts', 'Unit Price', 'Total']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')

        # Get billable usage
        highwater_data = self.query_helper.get_highwater_for_month(year, month, billable_only=True)

        row = 2
        for hw in highwater_data:
            domain = self.query_helper.session.query(Domain).get(hw.domain_id)
            if not domain:
                continue

            customer = self.query_helper.session.query(Customer).get(domain.customer_id)
            if not customer:
                continue

            cos_mapping = self.query_helper.get_cos_mapping_by_id(hw.cos_id)
            if not cos_mapping:
                continue

            # Fetch current price from QuickBooks
            unit_price = self._get_item_price(cos_mapping.qbo_item_id)
            total = hw.highwater_count * unit_price

            ws.cell(row=row, column=1, value=customer.customer_name)
            ws.cell(row=row, column=2, value=domain.domain_name)
            ws.cell(row=row, column=3, value=cos_mapping.cos_name)
            quota_cell = ws.cell(row=row, column=4, value=cos_mapping.quota_gb or '')
            quota_cell.alignment = Alignment(horizontal='center')
            accounts_cell = ws.cell(row=row, column=5, value=hw.highwater_count)
            accounts_cell.alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=6, value=f"${unit_price:,.2f}")
            ws.cell(row=row, column=7, value=f"${total:,.2f}")

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12

    def _create_customer_breakdown_sheet(self, wb: Workbook, year: int, month: int) -> None:
        """Create per-customer breakdown sheet.

        Args:
            wb: Workbook
            year: Year
            month: Month
        """
        ws = wb.create_sheet("Customer Breakdown")

        # Get all customers with usage
        from collections import defaultdict
        customer_data = defaultdict(list)

        highwater_data = self.query_helper.get_highwater_for_month(year, month, billable_only=True)

        for hw in highwater_data:
            domain = self.query_helper.session.query(Domain).get(hw.domain_id)
            if not domain:
                continue

            customer = self.query_helper.session.query(Customer).get(domain.customer_id)
            if not customer:
                continue

            cos_mapping = self.query_helper.get_cos_mapping_by_id(hw.cos_id)
            if not cos_mapping:
                continue

            # Fetch current price from QuickBooks
            unit_price = self._get_item_price(cos_mapping.qbo_item_id)

            customer_data[customer.customer_name].append({
                'domain': domain.domain_name,
                'cos': cos_mapping.cos_name,
                'quota_gb': cos_mapping.quota_gb,
                'count': hw.highwater_count,
                'unit_price': unit_price,
                'total': hw.highwater_count * unit_price
            })

        # Generate breakdown for each customer
        row = 1
        for customer_name in sorted(customer_data.keys()):
            # Customer header
            ws.cell(row=row, column=1, value=customer_name).font = Font(bold=True, size=12)
            row += 1

            # Column headers
            headers = ['Domain', 'CoS', 'Quota (GB)', 'Accounts', 'Unit Price', 'Total']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_num, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
            row += 1

            # Data rows
            customer_total = 0.0
            for item in customer_data[customer_name]:
                ws.cell(row=row, column=1, value=item['domain'])
                ws.cell(row=row, column=2, value=item['cos'])
                quota_cell = ws.cell(row=row, column=3, value=item['quota_gb'] or '')
                quota_cell.alignment = Alignment(horizontal='center')
                accounts_cell = ws.cell(row=row, column=4, value=item['count'])
                accounts_cell.alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=5, value=f"${item['unit_price']:,.2f}")
                ws.cell(row=row, column=6, value=f"${item['total']:,.2f}")
                customer_total += item['total']
                row += 1

            # Customer total
            ws.cell(row=row, column=5, value="TOTAL:").font = self.total_font
            ws.cell(row=row, column=6, value=f"${customer_total:,.2f}").font = self.total_font
            row += 2

        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12

    def _create_nonbillable_sheet(self, wb: Workbook, year: int, month: int) -> None:
        """Create sheet showing non-billable usage (excluded items).

        Args:
            wb: Workbook
            year: Year
            month: Month
        """
        ws = wb.create_sheet("Non-Billable Usage")

        # Headers
        headers = ['Domain', 'CoS', 'Accounts', 'Reason']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        # Get non-billable usage
        highwater_data = self.query_helper.session.query(MonthlyHighwater).filter(
            MonthlyHighwater.year == year,
            MonthlyHighwater.month == month,
            MonthlyHighwater.billable == False
        ).all()

        row = 2
        for hw in highwater_data:
            domain = self.query_helper.session.query(Domain).get(hw.domain_id)
            cos_mapping = self.query_helper.get_cos_mapping_by_id(hw.cos_id)

            if not domain or not cos_mapping:
                continue

            # Determine exclusion reason
            if self.query_helper.is_domain_excluded(domain.domain_name):
                reason = "Domain excluded"
            elif self.query_helper.is_cos_excluded(cos_mapping.cos_name):
                reason = "CoS excluded"
            else:
                reason = "Unknown"

            ws.cell(row=row, column=1, value=domain.domain_name)
            ws.cell(row=row, column=2, value=cos_mapping.cos_name)
            accounts_cell = ws.cell(row=row, column=3, value=hw.highwater_count)
            accounts_cell.alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=4, value=reason)
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20

    def _get_month_name(self, month: int) -> str:
        """Get month name from number.

        Args:
            month: Month number (1-12)

        Returns:
            Month name
        """
        months = [
            'January', 'February', 'March', 'April', 'May', 'June',
            'July', 'August', 'September', 'October', 'November', 'December'
        ]
        return months[month - 1] if 1 <= month <= 12 else str(month)


def generate_monthly_report(year: int, month: int, query_helper: QueryHelper,
                           output_path: Optional[str] = None, qbo_client=None) -> str:
    """Convenience function to generate monthly report.

    Args:
        year: Year
        month: Month
        query_helper: Database query helper
        output_path: Optional output path
        qbo_client: Optional QuickBooks client for fetching current prices

    Returns:
        Path to generated report
    """
    generator = ExcelReportGenerator(query_helper, qbo_client)
    return generator.generate_monthly_report(year, month, output_path)
